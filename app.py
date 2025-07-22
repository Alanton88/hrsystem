from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file
import gspread
from google.oauth2.service_account import Credentials
import os
from dotenv import load_dotenv
import json
from datetime import datetime
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
import sys
from google_auth_oauthlib.flow import InstalledAppFlow
from gas_integration import set_purchase_protection
import io
import pandas as pd
from googleapiclient.http import MediaIoBaseDownload
import openpyxl
from openpyxl.utils import get_column_letter
import requests
import re

# 載入環境變數
load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'your-secret-key-here')

# Railway 環境設定
if os.getenv('RAILWAY_ENVIRONMENT'):
    app.config['PREFERRED_URL_SCHEME'] = 'https'

# Google Sheets API 設定
SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive'
]

def get_google_sheets_client():
    """建立Google Sheets客戶端"""
    try:
        # 從環境變數或服務帳戶金鑰檔案讀取憑證
        service_account_info = os.getenv('GOOGLE_SERVICE_ACCOUNT_INFO')
        if service_account_info:
            creds_dict = json.loads(service_account_info)
        else:
            # 如果沒有環境變數，嘗試從檔案讀取
            creds_dict = json.load(open('service-account-key.json'))
        
        creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
        client = gspread.authorize(creds)
        return client
    except Exception as e:
        print(f"Google Sheets 客戶端建立失敗: {e}")
        return None

def verify_credentials(username, password):
    """驗證使用者帳號密碼"""
    try:
        client = get_google_sheets_client()
        if not client:
            return False, "無法連接到Google Sheets"
        
        # 開啟試算表（需要設定試算表ID）
        spreadsheet_id = os.getenv('SPREADSHEET_ID', '1ZB6ri0fzqTRk_ciHibcGXEuViNmW1Ag9kkazE8A5iKc')
        if not spreadsheet_id:
            return False, "未設定試算表ID"
        
        spreadsheet = client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet('使用者帳號')  # 工作表名稱
        
        # 取得所有資料
        all_records = worksheet.get_all_records()
        
        # 檢查帳號密碼
        for record in all_records:
            if record.get('帳號') == username and record.get('密碼') == password:
                return True, "登入成功"
        
        return False, "帳號或密碼錯誤"
        
    except Exception as e:
        print(f"驗證過程發生錯誤: {e}")
        return False, "系統錯誤，請稍後再試"

def get_dropdown_list(sheet_name, col_name):
    """從Google Sheets取得下拉選單資料"""
    try:
        client = get_google_sheets_client()
        if not client:
            return []
        spreadsheet_id = os.getenv('SPREADSHEET_ID', '1ZB6ri0fzqTRk_ciHibcGXEuViNmW1Ag9kkazE8A5iKc')
        if not spreadsheet_id:
            return []
        spreadsheet = client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet(sheet_name)
        records = worksheet.get_all_records()
        return [r[col_name] for r in records if col_name in r and r[col_name]]
    except Exception as e:
        print(f"取得下拉選單資料失敗: {e}")
        return []

def get_safe_records(worksheet):
    """安全地從工作表取得記錄，處理重複標題問題"""
    try:
        return worksheet.get_all_records()
    except Exception as e:
        if "duplicates" in str(e):
            print("DEBUG: 檢測到重複標題，使用手動標題設定")
            # 取得第一行作為標題
            headers = worksheet.row_values(1)
            # 清理重複和空標題
            cleaned_headers = []
            for i, header in enumerate(headers):
                if header:
                    cleaned_headers.append(header)
                else:
                    cleaned_headers.append(f'Column_{i+1}')
            
            # 取得所有資料行
            all_values = worksheet.get_all_values()
            if len(all_values) > 1:
                data_rows = all_values[1:]  # 跳過標題行
                all_records = []
                for row in data_rows:
                    record = {}
                    for i, value in enumerate(row):
                        if i < len(cleaned_headers):
                            record[cleaned_headers[i]] = value
                    all_records.append(record)
                return all_records
            else:
                return []
        else:
            raise e

def get_user_info(username):
    """依帳號取得姓名、mail與role"""
    try:
        client = get_google_sheets_client()
        if not client:
            return {'name': '', 'mail': '', 'role': ''}
        spreadsheet_id = os.getenv('SPREADSHEET_ID', '1ZB6ri0fzqTRk_ciHibcGXEuViNmW1Ag9kkazE8A5iKc')
        if not spreadsheet_id:
            return {'name': '', 'mail': '', 'role': ''}
        spreadsheet = client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet('使用者帳號')
        records = get_safe_records(worksheet)
        for r in records:
            if r.get('帳號') == username:
                return {
                    'name': r.get('姓名', ''),
                    'mail': r.get('mail', ''),
                    'role': r.get('角色', '')
                }
        return {'name': '', 'mail': '', 'role': ''}
    except Exception as e:
        print(f"取得user info失敗: {e}")
        return {'name': '', 'mail': '', 'role': ''}

def generate_purchase_no():
    """產生請購單號 (YYYYmmdd-流水號)，確保在請購單唯一且連號"""
    try:
        client = get_google_sheets_client()
        if not client:
            return datetime.now().strftime('%Y%m%d-001')
        spreadsheet_id = os.getenv('SPREADSHEET_ID')
        spreadsheet = client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet('請購單')
        today = datetime.now().strftime('%Y%m%d')
        records = worksheet.get_all_records()
        # 收集今天所有已用過的流水號
        seqs = []
        for r in records:
            no = str(r.get('請購單號', ''))
            if no.startswith(today):
                try:
                    seq = int(no.split('-')[1])
                    seqs.append(seq)
                except:
                    continue
        next_seq = max(seqs) + 1 if seqs else 1
        return f"{today}-{next_seq:03d}"
    except Exception as e:
        print(f"產生請購單號失敗: {e}")
        return datetime.now().strftime('%Y%m%d-001')

def upload_to_drive(file_storage, folder_id):
    print('=== Flask upload debug ===')
    print('folder_id:', folder_id)
    print('env folder_id:', os.getenv('PURCHASE_REQUEST_ATTACHMENT_FOLDER_ID'))

    try:
        # 使用 OAuth 2.0 流程進行檔案上傳
        SCOPES = ['https://www.googleapis.com/auth/drive.file']
        
        # Railway 環境下使用服務帳戶認證
        if os.getenv('RAILWAY_ENVIRONMENT'):
            # 使用環境變數中的服務帳戶資訊
            service_account_info = os.getenv('GOOGLE_SERVICE_ACCOUNT_INFO')
            if service_account_info:
                creds_dict = json.loads(service_account_info)
                creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
            else:
                raise Exception("Railway 環境下需要設定 GOOGLE_SERVICE_ACCOUNT_INFO")
        else:
            # 本地環境使用 OAuth 流程
            flow = InstalledAppFlow.from_client_secrets_file(
                'client_secret_530091699052-b5t9bbrevaodte3fi0u4n5urd9l6rhvs.apps.googleusercontent.com.json', 
                SCOPES
            )
            creds = flow.run_local_server(port=55551)
        service = build('drive', 'v3', credentials=creds)
        
        filename = file_storage.filename
        # 使用更安全的臨時檔案名稱，避免中文檔名問題
        import uuid
        temp_filename = f"tmp_{uuid.uuid4().hex}_{filename}"
        temp_path = temp_filename
        file_storage.save(temp_path)
        file_metadata = {
            'name': filename,
            'parents': [folder_id]
        }
        media = MediaFileUpload(temp_path, resumable=True)
        file = service.files().create(body=file_metadata, media_body=media, fields='id,webViewLink,webContentLink').execute()
        
        print('建立成功:', file)
        print('檔案ID:', file.get('id'))
        print('webViewLink:', file.get('webViewLink'))
        print('webContentLink:', file.get('webContentLink'))
        
        # 確保檔案可以被公開存取
        try:
            # 設定檔案權限為任何人都可以查看
            permission = {
                'type': 'anyone',
                'role': 'reader'
            }
            service.permissions().create(
                fileId=file.get('id'),
                body=permission,
                fields='id'
            ).execute()
            print('檔案權限設定成功')
        except Exception as perm_error:
            print(f'設定檔案權限失敗: {perm_error}')
        
        # 如果 webViewLink 為空，使用 webContentLink 或手動構建連結
        web_view_link = file.get('webViewLink')
        if not web_view_link:
            file_id = file.get('id')
            web_view_link = f"https://drive.google.com/file/d/{file_id}/view"
            print(f'手動構建連結: {web_view_link}')
        
        # 安全地刪除臨時檔案
        try:
            os.remove(temp_path)
            print(f'臨時檔案刪除成功: {temp_path}')
        except Exception as del_error:
            print(f'臨時檔案刪除失敗: {del_error}')
            # 不影響主要功能，繼續執行
        
        return web_view_link
    except Exception as e:
        import traceback
        print('上傳附件失敗:', e)
        traceback.print_exc()
        return ''

@app.route('/')
def index():
    """首頁"""
    if 'logged_in' in session and session['logged_in']:
        return redirect(url_for('dashboard'))
    return render_template('login.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    """登入處理"""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if not username or not password:
            flash('請輸入帳號和密碼', 'error')
            return render_template('login.html')
        
        success, message = verify_credentials(username, password)
        
        if success:
            session['logged_in'] = True
            session['username'] = username
            user_info = get_user_info(username)
            session['role'] = user_info.get('role', '')
            now = datetime.now().strftime('%Y%m%d %H:%M')
            write_system_log(user_info.get('name', username), login_time=now)
            flash(message, 'success')
            return redirect(url_for('dashboard'))
        else:
            flash(message, 'error')
            return render_template('login.html')
    
    return render_template('login.html')

@app.route('/dashboard')
def dashboard():
    """儀表板頁面"""
    if 'logged_in' not in session or not session['logged_in']:
        return redirect(url_for('index'))
    
    try:
        # 計算製造部門待簽核筆數
        client = get_google_sheets_client()
        spreadsheet_id = os.getenv('SPREADSHEET_ID')
        spreadsheet = client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet('請購單')
        
        # 取得所有資料
        all_records = get_safe_records(worksheet)
        
        # 篩選製造部門的請購單（申請部門不等於研發部門）
        manufacturing_records = []
        for record in all_records:
            dept = record.get('請購部門', '').strip()
            sign_status = record.get('請購單簽核', '').strip()
            if dept != '研發部' and sign_status == '待簽核':
                manufacturing_records.append(record)
        
        # 計算待簽核筆數
        pending_count = 0
        for record in manufacturing_records:
            if record.get('請購單簽核') == '待簽核' or not record.get('請購單簽核'):
                pending_count += 1
        
        # 計算研發部門待簽核筆數
        rd_pending_count = 0
        for record in all_records:
            if record.get('請購部門') == '研發部' and (record.get('請購單簽核') == '待簽核' or not record.get('請購單簽核')):
                rd_pending_count += 1
        
        return render_template('dashboard.html', 
                             username=session.get('username'),
                             manufacturing_pending_count=pending_count,
                             rd_pending_count=rd_pending_count)
    except Exception as e:
        print(f"取得待簽核筆數失敗: {e}")
        return render_template('dashboard.html', 
                             username=session.get('username'),
                             manufacturing_pending_count=0,
                             rd_pending_count=0)

@app.route('/logout')
def logout():
    if 'username' in session:
        now = datetime.now().strftime('%Y%m%d %H:%M')
        user_info = get_user_info(session['username'])
        update_system_log(user_info.get('name', session['username']), logout_time=now)
    session.clear()
    flash('已成功登出', 'success')
    return redirect(url_for('index'))

@app.route('/purchase-approval/<dept>')
def purchase_approval(dept):
    """請購單簽核頁面"""
    if 'logged_in' not in session or not session['logged_in']:
        return redirect(url_for('index'))
    if session.get('role') == '一般人員':
        return '權限不足，無法存取此頁面', 403
    
    # 根據部門參數設定頁面標題
    dept_names = {
        'manufacturing': '製造部門',
        'rd': '研發部門'
    }
    
    dept_name = dept_names.get(dept, '未知部門')
    
    try:
        # 從 Google Sheets 取得該部門的請購單資料
        client = get_google_sheets_client()
        spreadsheet_id = os.getenv('SPREADSHEET_ID')
        spreadsheet = client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet('請購單')
        
        # 取得所有資料
        all_records = get_safe_records(worksheet)
        
        # 篩選請購單
        filtered_records = []
        for record in all_records:
            # 只顯示狀態為「待簽核」或空的請購單
            approval_status = record.get('請購單簽核', '')
            if approval_status not in ['待簽核', '']:
                continue  # 跳過已核准或駁回的請購單
            
            if dept == 'manufacturing':
                # 製造部門：篩選申請部門不等於研發部門的請購單
                if record.get('請購部門') != '研發部':
                    filtered_records.append(record)
            elif dept == 'rd':
                # 研發部門：篩選申請部門等於研發部門的請購單
                if record.get('請購部門') == '研發部':
                    filtered_records.append(record)
        
        return render_template('purchase_approval.html', 
                             records=filtered_records, 
                             dept_name=dept_name,
                             dept_code=dept)
        
    except Exception as e:
        print(f"取得請購單資料失敗: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/get-manufacturing-pending-count')
def get_manufacturing_pending_count():
    """獲取製造部門待簽核筆數"""
    if 'logged_in' not in session or not session['logged_in']:
        return jsonify({'success': False, 'message': '未登入'})
    
    try:
        client = get_google_sheets_client()
        spreadsheet_id = os.getenv('SPREADSHEET_ID')
        spreadsheet = client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet('請購單')
        
        # 取得所有資料
        all_records = get_safe_records(worksheet)
        
        # 篩選製造部門的請購單（申請部門不等於研發部門）
        manufacturing_records = []
        for record in all_records:
            dept = record.get('請購部門', '').strip()
            sign_status = record.get('請購單簽核', '').strip()
            if dept != '研發部' and sign_status == '待簽核':
                manufacturing_records.append(record)
        
        # 計算待簽核筆數
        pending_count = 0
        for record in manufacturing_records:
            if record.get('請購單簽核') == '待簽核' or not record.get('請購單簽核'):
                pending_count += 1
        
        return jsonify({'success': True, 'count': pending_count})
        
    except Exception as e:
        print(f"取得待簽核筆數失敗: {e}")
        return jsonify({'success': False, 'message': f'取得失敗: {str(e)}'})

@app.route('/get-rd-pending-count')
def get_rd_pending_count():
    """獲取研發部門待簽核筆數"""
    if 'logged_in' not in session or not session['logged_in']:
        return jsonify({'success': False, 'message': '未登入'})
    
    try:
        client = get_google_sheets_client()
        spreadsheet_id = os.getenv('SPREADSHEET_ID')
        spreadsheet = client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet('請購單')
        
        # 取得所有資料
        all_records = get_safe_records(worksheet)
        
        # 篩選研發部門的請購單（申請部門等於研發部門）
        rd_records = [r for r in all_records if r.get('請購部門') == '研發部']
        
        # 計算待簽核筆數
        pending_count = 0
        for record in rd_records:
            if record.get('請購單簽核') == '待簽核' or not record.get('請購單簽核'):
                pending_count += 1
        
        return jsonify({'success': True, 'count': pending_count})
        
    except Exception as e:
        print(f"取得研發部門待簽核筆數失敗: {e}")
        return jsonify({'success': False, 'message': f'取得失敗: {str(e)}'})

@app.route('/purchase-detail/<purchase_no>')
def purchase_detail(purchase_no):
    """取得單一請購單詳細資料"""
    if 'logged_in' not in session or not session['logged_in']:
        return jsonify({'success': False, 'message': '未登入'})
    
    try:
        client = get_google_sheets_client()
        spreadsheet_id = os.getenv('SPREADSHEET_ID')
        spreadsheet = client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet('請購單')
        
        # 取得所有資料
        all_records = get_safe_records(worksheet)
        
        # 找到對應的請購單
        purchase_record = None
        for record in all_records:
            if str(record.get('請購單號', '')).replace('-', '') == purchase_no:
                purchase_record = record
                break
        
        if not purchase_record:
            return jsonify({'success': False, 'message': '找不到請購單'})
        
        return jsonify({'success': True, 'data': purchase_record})
        
    except Exception as e:
        print(f"取得請購單詳細資料失敗: {e}")
        return jsonify({'success': False, 'message': f'取得失敗: {str(e)}'})

@app.route('/receipt-management')
def receipt_management():
    """驗收單作業頁面"""
    if 'logged_in' not in session or not session['logged_in']:
        return redirect(url_for('index'))
    
    try:
        # 從 Google Sheets 取得已核准的請購單資料
        client = get_google_sheets_client()
        if not client:
            raise Exception("無法建立 Google Sheets 客戶端")
        
        spreadsheet_id = os.getenv('SPREADSHEET_ID', '1ZB6ri0fzqTRk_ciHibcGXEuViNmW1Ag9kkazE8A5iKc')
        if not spreadsheet_id:
            raise Exception("未設定試算表 ID")
        
        spreadsheet = client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet('請購單')
        
        # 取得所有資料
        all_records = get_safe_records(worksheet)
        
        # 篩選已核准的請購單
        approved_records = []
        departments = set()  # 用於收集部門列表
        applicants = set()   # 用於收集申請人列表
        
        print(f"DEBUG: 總記錄數: {len(all_records)}")
        if all_records:
            print(f"DEBUG: 第一筆記錄的欄位: {list(all_records[0].keys())}")
            print(f"DEBUG: 第一筆記錄: {all_records[0]}")
        
        for record in all_records:
            # 檢查多個可能的簽核欄位名稱
            approval_status = (
                record.get('請購單簽核') or 
                record.get('簽核狀態') or 
                record.get('approval_status') or 
                record.get('狀態') or 
                ''
            )
            
            print(f"DEBUG: 請購單號 {record.get('請購單號', 'N/A')} 的簽核狀態: '{approval_status}'")
            
            # 檢查多種可能的核准狀態值
            if approval_status in ['核准', 'approved', 'APPROVED', 'Approved', '已核准', '已批准']:
                # 為每個已核准的請購單設定預設驗收狀態（嘗試多個可能的欄位名稱）
                receipt_status = (
                    record.get('驗收單狀態') or 
                    record.get('驗收狀態') or 
                    record.get('receipt_status') or 
                    '待驗收'
                )
                record['驗收狀態'] = receipt_status  # 統一使用 '驗收狀態' 作為顯示欄位
                
                # 檢查編輯狀態欄位
                edit_status = record.get('編輯狀態', '可編輯')
                record['編輯狀態'] = edit_status
                
                # 只顯示未設為唯讀的記錄
                if edit_status != '唯讀':
                    approved_records.append(record)
                    print(f"DEBUG: 找到已核准請購單: {record.get('請購單號')}，編輯狀態: {edit_status}")
                    
                    # 收集部門和申請人
                    if record.get('請購部門'):
                        departments.add(record.get('請購部門'))
                    if record.get('申請人'):
                        applicants.add(record.get('申請人'))
                else:
                    print(f"DEBUG: 跳過唯讀請購單: {record.get('請購單號')}")
        
        print(f"DEBUG: 找到 {len(approved_records)} 筆已核准請購單")
        
        # 轉換為排序列表
        departments = sorted(list(departments))
        applicants = sorted(list(applicants))
        
        # 取得登入者資訊
        username = session.get('username', '')
        user_info = get_user_info(username)
        
        return render_template('receipt_management.html', 
                             approved_records=approved_records,
                             total_approved=len(approved_records),
                             departments=departments,
                             applicants=applicants,
                             current_user_name=user_info.get('name', username))
        
    except Exception as e:
        print(f"取得驗收單資料失敗: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/purchase-summary')
def purchase_summary():
    """請購單彙總為採購清單頁面"""
    if 'logged_in' not in session or not session['logged_in']:
        return redirect(url_for('index'))
    
    try:
        # 從 Google Sheets 取得已核准的請購單資料
        client = get_google_sheets_client()
        if not client:
            raise Exception("無法建立 Google Sheets 客戶端")
        
        spreadsheet_id = os.getenv('SPREADSHEET_ID', '1ZB6ri0fzqTRk_ciHibcGXEuViNmW1Ag9kkazE8A5iKc')
        if not spreadsheet_id:
            raise Exception("未設定試算表 ID")
        
        spreadsheet = client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet('請購單')
        
        # 取得所有資料
        all_records = get_safe_records(worksheet)
        
        # 篩選已核准的請購單
        approved_records = []
        for record in all_records:
            if record.get('請購單簽核') == '核准':
                approved_records.append(record)
        
        # 按品名分組彙總
        summary_data = {}
        for record in approved_records:
            item_name = record.get('品名', '')
            spec = record.get('規格', '')
            unit = record.get('單位', '')
            
            # 使用品名+規格作為唯一鍵
            key = f"{item_name}_{spec}_{unit}"
            
            if key not in summary_data:
                summary_data[key] = {
                    '品名': item_name,
                    '規格': spec,
                    '單位': unit,
                    '總數量': 0,
                    '請購單號列表': [],
                    '申請部門列表': [],
                    '需求日期列表': []
                }
            
            # 累加數量
            try:
                quantity = int(record.get('數量', 0))
                summary_data[key]['總數量'] += quantity
            except (ValueError, TypeError):
                summary_data[key]['總數量'] += 0
            
            # 記錄請購單號
            purchase_no = record.get('請購單號', '')
            if purchase_no:
                summary_data[key]['請購單號列表'].append(purchase_no)
            
            # 記錄申請部門
            department = record.get('請購部門', '')
            if department:
                summary_data[key]['申請部門列表'].append(department)
            
            # 記錄需求日期
            need_date = record.get('需求日期', '')
            if need_date:
                summary_data[key]['需求日期列表'].append(need_date)
        
        # 轉換為列表格式
        summary_list = list(summary_data.values())
        
        return render_template('purchase_summary.html', 
                             summary_data=summary_list,
                             total_items=len(summary_list),
                             total_orders=len(approved_records))
        
    except Exception as e:
        print(f"取得採購清單失敗: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/update-approval-status', methods=['POST'])
def update_approval_status():
    """更新請購單簽核狀態"""
    if 'logged_in' not in session or not session['logged_in']:
        return jsonify({'success': False, 'message': '未登入'})
    
    try:
        data = request.get_json()
        purchase_no = data.get('purchase_no')
        status = data.get('status')
        reason = data.get('reason', '')
        
        if not purchase_no or not status:
            return jsonify({'success': False, 'message': '缺少必要參數'})
        
        # 取得當前登入者資訊
        username = session.get('username', '')
        user_info = get_user_info(username)
        current_user_name = user_info.get('name', username)
        print(f"DEBUG: 當前登入者: {current_user_name}")
        
        # 更新 Google Sheets 中的簽核狀態
        client = get_google_sheets_client()
        spreadsheet_id = os.getenv('SPREADSHEET_ID')
        spreadsheet = client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet('請購單')
        
        # 取得欄位標題
        headers = worksheet.row_values(1)
        print(f"DEBUG: 欄位標題: {headers}")
        
        # 找到對應的請購單號
        all_records = get_safe_records(worksheet)
        row_index = None
        
        for i, record in enumerate(all_records, start=2):  # start=2 因為第1行是標題
            if record.get('請購單號') == purchase_no:
                row_index = i
                break
        
        if row_index is None:
            return jsonify({'success': False, 'message': '找不到請購單'})
        
        # 找到相關欄位的索引
        approval_status_col = None
        approval_person_col = None
        approval_date_col = None
        reject_reason_col = None
        
        for i, header in enumerate(headers):
            if header == '請購單簽核':
                approval_status_col = i + 1  # gspread 使用 1-based 索引
            elif header == '請購單簽核人員':
                approval_person_col = i + 1
            elif header == '請購單簽核日期':
                approval_date_col = i + 1
            elif header == '駁回原因說明':
                reject_reason_col = i + 1
        
        print(f"DEBUG: 找到簽核狀態欄位 '{headers[approval_status_col-1] if approval_status_col else 'None'}' 在第 {approval_status_col} 欄")
        print(f"DEBUG: 找到請購單簽核人員欄位 '{headers[approval_person_col-1] if approval_person_col else 'None'}' 在第 {approval_person_col} 欄")
        print(f"DEBUG: 找到請購單簽核日期欄位 '{headers[approval_date_col-1] if approval_date_col else 'None'}' 在第 {approval_date_col} 欄")
        print(f"DEBUG: 找到請購單駁回原因欄位 '{headers[reject_reason_col-1] if reject_reason_col else 'None'}' 在第 {reject_reason_col} 欄")
        
        # 準備更新的資料
        updates = []
        
        # 更新簽核狀態
        if approval_status_col:
            updates.append((row_index, approval_status_col, status))
        
        # 更新簽核人員（只有當狀態為核准或駁回時才更新）
        if approval_person_col and status in ['核准', '駁回']:
            updates.append((row_index, approval_person_col, current_user_name))
        
        # 更新簽核日期（只有當狀態為核准或駁回時才更新）
        if approval_date_col and status in ['核准', '駁回']:
            current_date = datetime.now().strftime('%Y%m%d')
            updates.append((row_index, approval_date_col, current_date))
        
        # 更新請購單駁回原因（只有當狀態為駁回時才更新）
        if reject_reason_col and status == '駁回':
            updates.append((row_index, reject_reason_col, reason))
        
        print(f"DEBUG: 準備更新 {len(updates)} 個欄位")
        
        # 執行更新
        for row, col, value in updates:
            worksheet.update_cell(row, col, value)
            print(f"DEBUG: 已更新第 {row} 行第 {col} 欄為: {value}")
        
        return jsonify({
            'success': True, 
            'message': '簽核狀態已更新',
            'approver': current_user_name,
            'approval_date': datetime.now().strftime('%Y%m%d') if status in ['核准', '駁回'] else '',
            'approval_note': reason if status == '駁回' else ''
        })
        
    except Exception as e:
        print(f"更新簽核狀態失敗: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/update-receipt-approval', methods=['POST'])
def update_receipt_approval():
    """更新驗收簽核狀態"""
    if 'logged_in' not in session or not session['logged_in']:
        return jsonify({'success': False, 'message': '未登入'})
    
    try:
        data = request.get_json()
        purchase_no = data.get('purchase_no')
        approval_status = data.get('approval_status')
        approval_date = data.get('approval_date')
        approver = data.get('approver')
        approval_note = data.get('approval_note', '')
        
        if not purchase_no or not approval_status:
            return jsonify({'success': False, 'message': '缺少必要參數'})
        
        print(f"DEBUG: 更新驗收簽核狀態 - 請購單號: {purchase_no}, 狀態: {approval_status}")
        
        # 取得當前登入者資訊
        username = session.get('username', '')
        user_info = get_user_info(username)
        current_user_name = user_info.get('name', username)
        print(f"DEBUG: 當前登入者: {current_user_name}")
        
        # 更新 Google Sheets
        client = get_google_sheets_client()
        spreadsheet_id = os.getenv('SPREADSHEET_ID')
        spreadsheet = client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet('請購單')
        
        # 取得所有資料
        all_records = get_safe_records(worksheet)
        
        # 找到對應的請購單
        target_row = None
        for i, record in enumerate(all_records, start=2):  # 從第2行開始（跳過標題行）
            if str(record.get('請購單號', '')).replace('-', '') == purchase_no:
                target_row = i
                break
        
        if not target_row:
            return jsonify({'success': False, 'message': '找不到指定的請購單'})
        
        # 取得欄位標題
        headers = worksheet.row_values(1)
        print(f"DEBUG: 欄位標題: {headers}")
        
        # 找到驗收簽核相關欄位的索引
        approval_person_col = None
        approval_date_col = None
        approval_status_col = None
        approval_note_col = None
        
        for i, header in enumerate(headers):
            if header == '驗收簽核人員':
                approval_person_col = i + 1  # gspread 使用 1-based 索引
            elif header == '驗收簽核日期':
                approval_date_col = i + 1
            elif header == '驗收簽核狀態':
                approval_status_col = i + 1
            elif header == '驗收簽核備註':
                approval_note_col = i + 1
        
        print(f"DEBUG: 找到驗收簽核人員欄位 '{headers[approval_person_col-1] if approval_person_col else 'None'}' 在第 {approval_person_col} 欄")
        print(f"DEBUG: 找到驗收簽核日期欄位 '{headers[approval_date_col-1] if approval_date_col else 'None'}' 在第 {approval_date_col} 欄")
        print(f"DEBUG: 找到驗收簽核狀態欄位 '{headers[approval_status_col-1] if approval_status_col else 'None'}' 在第 {approval_status_col} 欄")
        print(f"DEBUG: 找到驗收簽核備註欄位 '{headers[approval_note_col-1] if approval_note_col else 'None'}' 在第 {approval_note_col} 欄")
        
        # 準備更新的資料
        updates = []
        
        if approval_person_col:
            updates.append((target_row, approval_person_col, current_user_name))
        if approval_date_col:
            updates.append((target_row, approval_date_col, approval_date))
        if approval_status_col:
            updates.append((target_row, approval_status_col, approval_status))
        if approval_note_col:
            updates.append((target_row, approval_note_col, approval_note))
        
        print(f"DEBUG: 更新欄位 - 簽核狀態: {approval_status_col}, 簽核日期: {approval_date_col}, 簽核人員: {approval_person_col}, 簽核備註: {approval_note_col}")
        
        # 執行更新
        for row, col, value in updates:
            worksheet.update_cell(row, col, value)
        
        print(f"DEBUG: 已更新 - 簽核狀態: {approval_status}, 簽核日期: {approval_date}, 簽核人員: {current_user_name}")
        
        # 檢查是否需要將記錄設為唯讀狀態
        # 條件：驗收簽核狀態為核准或駁回（已完成驗收簽核）
        if approval_status in ['核准', '駁回']:
            # 先執行原本的欄位標記（如有需要）
            # ...（原本的編輯狀態欄位處理）

            # 新增：呼叫 GAS 進行真正的 Google Sheets 欄位保護
            if os.getenv('ENABLE_GAS_PROTECTION', 'false').lower() == 'true':
                protection_result = set_purchase_protection(purchase_no, "驗收簽核完成")
                print(f"GAS 保護結果: {protection_result}")
        
        return jsonify({
            'success': True, 
            'message': '驗收簽核狀態已更新',
            'approver': current_user_name,
            'approval_date': approval_date,
            'approval_note': approval_note,
            'is_readonly': approval_status in ['核准', '駁回']
        })
        
    except Exception as e:
        print(f"更新驗收簽核狀態失敗: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/verify-admin-password', methods=['POST'])
def verify_admin_password():
    """驗證 admin 密碼並重新啟用編輯功能"""
    if 'logged_in' not in session or not session['logged_in']:
        return jsonify({'success': False, 'message': '未登入'})
    
    try:
        data = request.get_json()
        password = data.get('password')
        purchase_no = data.get('purchase_no')
        
        if not password or not purchase_no:
            return jsonify({'success': False, 'message': '缺少必要參數'})
        
        # 驗證 admin 密碼
        admin_username = 'admin'
        if verify_credentials(admin_username, password):
            # 密碼正確，重新啟用編輯功能
            client = get_google_sheets_client()
            spreadsheet_id = os.getenv('SPREADSHEET_ID')
            spreadsheet = client.open_by_key(spreadsheet_id)
            worksheet = spreadsheet.worksheet('請購單')
            
            # 取得欄位標題
            headers = worksheet.row_values(1)
            
            # 找到編輯狀態欄位的索引（使用現有的最後一欄）
            edit_status_col = None
            for i, header in enumerate(headers):
                if header == '編輯狀態':
                    edit_status_col = i + 1
                    break
            
            if not edit_status_col:
                # 檢查是否有空欄位可以使用
                empty_col = None
                for i, header in enumerate(headers):
                    if not header or header.strip() == '':
                        empty_col = i + 1
                        break
                
                if empty_col:
                    # 使用空欄位
                    edit_status_col = empty_col
                    worksheet.update_cell(1, edit_status_col, '編輯狀態')
                    print(f"DEBUG: 使用空欄位第 {edit_status_col} 欄作為編輯狀態欄位")
                else:
                    # 檢查是否可以在現有範圍內添加新欄位
                    current_cols = len(headers)
                    if current_cols < 24:  # Google Sheets 最大欄位數
                        edit_status_col = current_cols + 1
                        worksheet.update_cell(1, edit_status_col, '編輯狀態')
                        print(f"DEBUG: 新增編輯狀態欄位在第 {edit_status_col} 欄")
                    else:
                        # 如果無法添加新欄位，使用最後一欄覆蓋
                        edit_status_col = 24
                        worksheet.update_cell(1, edit_status_col, '編輯狀態')
                        print(f"DEBUG: 使用最後一欄第 {edit_status_col} 欄作為編輯狀態欄位")
            
            if purchase_no == 'ALL':
                # 重新啟用所有記錄的編輯功能
                all_records = get_safe_records(worksheet)
                updated_count = 0
                
                for i, record in enumerate(all_records, start=2):
                    # 檢查當前編輯狀態
                    current_status = worksheet.cell(i, edit_status_col).value
                    if current_status == '唯讀':
                        worksheet.update_cell(i, edit_status_col, '可編輯')
                        updated_count += 1
                
                # 嘗試解除所有 Google Sheets 保護
                try:
                    # 使用 Google Sheets API 解除保護
                    from googleapiclient.discovery import build
                    from googleapiclient.errors import HttpError
                    
                    # 取得認證憑證
                    creds = None
                    if os.path.exists('token.json'):
                        from google.auth.transport.requests import Request
                        from google.oauth2.credentials import Credentials
                        creds = Credentials.from_authorized_user_file('token.json', ['https://www.googleapis.com/auth/spreadsheets'])
                    elif os.path.exists('service-account-key.json'):
                        from google.oauth2 import service_account
                        creds = service_account.Credentials.from_service_account_file(
                            'service-account-key.json',
                            scopes=['https://www.googleapis.com/auth/spreadsheets']
                        )
                    
                    if creds and creds.valid:
                        service = build('sheets', 'v4', credentials=creds)
                        
                        # 取得所有保護範圍
                        spreadsheet = service.spreadsheets().get(
                            spreadsheetId=spreadsheet_id,
                            ranges=[],
                            includeGridData=False
                        ).execute()
                        
                        # 刪除所有保護範圍
                        requests = []
                        if 'sheets' in spreadsheet:
                            for sheet in spreadsheet['sheets']:
                                if 'protectedRanges' in sheet:
                                    for protected_range in sheet['protectedRanges']:
                                        requests.append({
                                            'deleteProtectedRange': {
                                                'protectedRangeId': protected_range['protectedRangeId']
                                            }
                                        })
                        
                        if requests:
                            body = {'requests': requests}
                            result = service.spreadsheets().batchUpdate(
                                spreadsheetId=spreadsheet_id,
                                body=body
                            ).execute()
                            print(f"DEBUG: 已成功解除所有 Google Sheets 保護")
                        else:
                            print(f"DEBUG: 沒有找到需要解除的保護範圍")
                            
                    else:
                        print(f"DEBUG: 無法取得有效的認證憑證，跳過 Google Sheets 保護解除")
                            
                except Exception as e:
                    print(f"DEBUG: 解除保護失敗: {e}")
                    print(f"DEBUG: 將使用應用程式層面的編輯控制")
                
                return jsonify({
                    'success': True, 
                    'message': f'密碼驗證成功，已重新啟用 {updated_count} 筆記錄的編輯功能',
                    'can_edit': True,
                    'updated_count': updated_count
                })
            else:
                # 重新啟用特定請購單的編輯功能
                all_records = get_safe_records(worksheet)
                
                # 找到對應的請購單
                target_row = None
                for i, record in enumerate(all_records, start=2):
                    if str(record.get('請購單號', '')).replace('-', '') == purchase_no:
                        target_row = i
                        break
                
                if not target_row:
                    return jsonify({'success': False, 'message': '找不到指定的請購單'})
                
                # 設定為可編輯
                worksheet.update_cell(target_row, edit_status_col, '可編輯')
                
                # 嘗試解除 Google Sheets 保護
                try:
                    # 使用 Google Sheets API 解除保護
                    from googleapiclient.discovery import build
                    from googleapiclient.errors import HttpError
                    
                    # 取得認證憑證
                    creds = None
                    if os.path.exists('token.json'):
                        from google.auth.transport.requests import Request
                        from google.oauth2.credentials import Credentials
                        creds = Credentials.from_authorized_user_file('token.json', ['https://www.googleapis.com/auth/spreadsheets'])
                    elif os.path.exists('service-account-key.json'):
                        from google.oauth2 import service_account
                        creds = service_account.Credentials.from_service_account_file(
                            'service-account-key.json',
                            scopes=['https://www.googleapis.com/auth/spreadsheets']
                        )
                    
                    if creds and creds.valid:
                        service = build('sheets', 'v4', credentials=creds)
                        
                        # 取得所有保護範圍
                        spreadsheet = service.spreadsheets().get(
                            spreadsheetId=spreadsheet_id,
                            ranges=[],
                            includeGridData=False
                        ).execute()
                        
                        # 找到並刪除特定請購單的保護範圍
                        requests = []
                        if 'sheets' in spreadsheet:
                            for sheet in spreadsheet['sheets']:
                                if 'protectedRanges' in sheet:
                                    for protected_range in sheet['protectedRanges']:
                                        # 檢查保護範圍的描述是否包含請購單號
                                        if 'description' in protected_range and purchase_no in protected_range['description']:
                                            requests.append({
                                                'deleteProtectedRange': {
                                                    'protectedRangeId': protected_range['protectedRangeId']
                                                }
                                            })
                        
                        if requests:
                            body = {'requests': requests}
                            result = service.spreadsheets().batchUpdate(
                                spreadsheetId=spreadsheet_id,
                                body=body
                            ).execute()
                            print(f"DEBUG: 已成功解除請購單 {purchase_no} 的 Google Sheets 保護")
                        else:
                            print(f"DEBUG: 沒有找到請購單 {purchase_no} 的保護範圍")
                            
                    else:
                        print(f"DEBUG: 無法取得有效的認證憑證，跳過 Google Sheets 保護解除")
                            
                except Exception as e:
                    print(f"DEBUG: 解除保護失敗: {e}")
                    print(f"DEBUG: 將使用應用程式層面的編輯控制")
                
                return jsonify({
                    'success': True, 
                    'message': '密碼驗證成功，已重新啟用編輯功能',
                    'can_edit': True
                })
        else:
            return jsonify({'success': False, 'message': '密碼錯誤'})
        
    except Exception as e:
        print(f"驗證密碼失敗: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/lock-record', methods=['POST'])
def lock_record():
    """鎖定記錄，設定為唯讀狀態"""
    if 'logged_in' not in session or not session['logged_in']:
        return jsonify({'success': False, 'message': '未登入'})
    
    try:
        data = request.get_json()
        purchase_no = data.get('purchase_no')
        
        print(f"DEBUG: 鎖定記錄 - 請購單號: {purchase_no}")
        
        if not purchase_no:
            return jsonify({'success': False, 'message': '缺少必要參數'})
        
        # 更新 Google Sheets，設定為唯讀狀態
        client = get_google_sheets_client()
        spreadsheet_id = os.getenv('SPREADSHEET_ID')
        spreadsheet = client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet('請購單')
        
        # 取得所有資料
        all_records = get_safe_records(worksheet)
        
        # 找到對應的請購單
        target_row = None
        for i, record in enumerate(all_records, start=2):
            if str(record.get('請購單號', '')).replace('-', '') == purchase_no:
                target_row = i
                break
        
        if not target_row:
            print(f"DEBUG: 找不到請購單 {purchase_no}")
            return jsonify({'success': False, 'message': '找不到指定的請購單'})
        
        print(f"DEBUG: 找到請購單 {purchase_no} 在第 {target_row} 行")
        
        # 取得欄位標題
        headers = worksheet.row_values(1)
        print(f"DEBUG: 欄位標題: {headers}")
        
        # 找到編輯狀態欄位的索引（使用現有的最後一欄）
        edit_status_col = None
        for i, header in enumerate(headers):
            if header == '編輯狀態':
                edit_status_col = i + 1
                break
        
        if not edit_status_col:
            # 檢查是否有空欄位可以使用
            empty_col = None
            for i, header in enumerate(headers):
                if not header or header.strip() == '':
                    empty_col = i + 1
                    break
            
            if empty_col:
                # 使用空欄位
                edit_status_col = empty_col
                worksheet.update_cell(1, edit_status_col, '編輯狀態')
                print(f"DEBUG: 使用空欄位第 {edit_status_col} 欄作為編輯狀態欄位")
            else:
                # 檢查是否可以在現有範圍內添加新欄位
                current_cols = len(headers)
                if current_cols < 24:  # Google Sheets 最大欄位數
                    edit_status_col = current_cols + 1
                    worksheet.update_cell(1, edit_status_col, '編輯狀態')
                    print(f"DEBUG: 新增編輯狀態欄位在第 {edit_status_col} 欄")
                else:
                    # 如果無法添加新欄位，使用最後一欄覆蓋
                    edit_status_col = 24
                    worksheet.update_cell(1, edit_status_col, '編輯狀態')
                    print(f"DEBUG: 使用最後一欄第 {edit_status_col} 欄作為編輯狀態欄位")
        else:
            print(f"DEBUG: 找到編輯狀態欄位在第 {edit_status_col} 欄")
        
        # 使用編輯狀態欄位來標記鎖定狀態
        # 注意：由於 gspread 6.2.1 的限制，我們無法直接設定 Google Sheets 保護
        # 但我們可以在應用程式層面實現鎖定功能
        worksheet.update_cell(target_row, edit_status_col, '唯讀')
        print(f"DEBUG: 已將請購單 {purchase_no} 設定為唯讀狀態")
        
        return jsonify({
            'success': True, 
            'message': f'請購單 {purchase_no} 已鎖定為唯讀狀態',
            'is_locked': True
        })
        
    except Exception as e:
        print(f"鎖定記錄失敗: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/update-receipt-status', methods=['POST'])
def update_receipt_status():
    """更新驗收單驗收狀態"""
    if 'logged_in' not in session or not session['logged_in']:
        return jsonify({'success': False, 'message': '未登入'})
    
    try:
        data = request.get_json()
        purchase_no = data.get('purchase_no')
        receipt_status = data.get('receipt_status')  # '待驗收', '驗收中', '已驗收', '驗收異常'
        
        if not purchase_no or not receipt_status:
            return jsonify({'success': False, 'message': '缺少必要參數'})
        
        print(f"DEBUG: 更新驗收單驗收狀態 - 請購單號: {purchase_no}, 狀態: {receipt_status}")
        
        # 取得當前登入者資訊
        username = session.get('username', '')
        user_info = get_user_info(username)
        current_user_name = user_info.get('name', username)
        
        print(f"DEBUG: 當前登入者: {current_user_name}")
        
        # 更新 Google Sheets 中的驗收單驗收狀態
        client = get_google_sheets_client()
        spreadsheet_id = os.getenv('SPREADSHEET_ID')
        spreadsheet = client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet('請購單')
        
        # 找到對應的請購單號
        all_records = get_safe_records(worksheet)
        row_index = None
        
        for i, record in enumerate(all_records, start=2):  # start=2 因為第1行是標題
            if str(record.get('請購單號', '')).replace('-', '') == purchase_no:
                row_index = i
                break
        
        if row_index is None:
            return jsonify({'success': False, 'message': '找不到請購單'})
        
        # 先取得欄位標題來確定正確的欄位位置
        headers = worksheet.row_values(1)
        print(f"DEBUG: 欄位標題: {headers}")
        
        # 尋找驗收單驗收狀態欄位
        receipt_status_col = None
        receipt_person_col = None
        receipt_date_col = None
        
        for i, header in enumerate(headers, 1):
            if '驗收單狀態' in header or '驗收狀態' in header or 'receipt_status' in header:
                receipt_status_col = i
                print(f"DEBUG: 找到驗收狀態欄位 '{header}' 在第 {i} 欄")
            elif '驗收人員' in header or 'receipt_person' in header:
                receipt_person_col = i
                print(f"DEBUG: 找到驗收人員欄位 '{header}' 在第 {i} 欄")
            elif '驗收日期' in header or 'receipt_date' in header:
                receipt_date_col = i
                print(f"DEBUG: 找到驗收日期欄位 '{header}' 在第 {i} 欄")
        
        # 如果找不到驗收單驗收狀態欄位，尋找空欄位或添加新欄位
        if receipt_status_col is None:
            # 檢查是否有空欄位可以使用
            empty_col = None
            for i, header in enumerate(headers):
                if not header or header.strip() == '':
                    empty_col = i + 1
                    break
            
            if empty_col:
                receipt_status_col = empty_col
                worksheet.update_cell(1, receipt_status_col, '驗收單狀態')
                print(f"DEBUG: 使用空欄位第 {receipt_status_col} 欄作為驗收單狀態欄位")
            else:
                # 檢查是否可以在現有範圍內添加新欄位
                current_cols = len(headers)
                if current_cols < 24:  # Google Sheets 最大欄位數
                    receipt_status_col = current_cols + 1
                    worksheet.update_cell(1, receipt_status_col, '驗收單狀態')
                    print(f"DEBUG: 新增驗收單狀態欄位在第 {receipt_status_col} 欄")
                else:
                    # 如果無法添加新欄位，使用最後一欄覆蓋
                    receipt_status_col = 24
                    worksheet.update_cell(1, receipt_status_col, '驗收單狀態')
                    print(f"DEBUG: 使用最後一欄第 {receipt_status_col} 欄作為驗收單狀態欄位")
        
        # 如果找不到驗收人員欄位，尋找空欄位或添加新欄位
        if receipt_person_col is None:
            # 檢查是否有空欄位可以使用
            empty_col = None
            for i, header in enumerate(headers):
                if not header or header.strip() == '':
                    empty_col = i + 1
                    break
            
            if empty_col:
                receipt_person_col = empty_col
                worksheet.update_cell(1, receipt_person_col, '驗收人員')
                print(f"DEBUG: 使用空欄位第 {receipt_person_col} 欄作為驗收人員欄位")
            else:
                # 檢查是否可以在現有範圍內添加新欄位
                current_cols = len(headers)
                if current_cols < 24:  # Google Sheets 最大欄位數
                    receipt_person_col = current_cols + 1
                    worksheet.update_cell(1, receipt_person_col, '驗收人員')
                    print(f"DEBUG: 新增驗收人員欄位在第 {receipt_person_col} 欄")
                else:
                    # 如果無法添加新欄位，使用最後一欄覆蓋
                    receipt_person_col = 24
                    worksheet.update_cell(1, receipt_person_col, '驗收人員')
                    print(f"DEBUG: 使用最後一欄第 {receipt_person_col} 欄作為驗收人員欄位")
        
        # 如果找不到驗收日期欄位，尋找空欄位或添加新欄位
        if receipt_date_col is None:
            # 檢查是否有空欄位可以使用
            empty_col = None
            for i, header in enumerate(headers):
                if not header or header.strip() == '':
                    empty_col = i + 1
                    break
            
            if empty_col:
                receipt_date_col = empty_col
                worksheet.update_cell(1, receipt_date_col, '驗收日期')
                print(f"DEBUG: 使用空欄位第 {receipt_date_col} 欄作為驗收日期欄位")
            else:
                # 檢查是否可以在現有範圍內添加新欄位
                current_cols = len(headers)
                if current_cols < 24:  # Google Sheets 最大欄位數
                    receipt_date_col = current_cols + 1
                    worksheet.update_cell(1, receipt_date_col, '驗收日期')
                    print(f"DEBUG: 新增驗收日期欄位在第 {receipt_date_col} 欄")
                else:
                    # 如果無法添加新欄位，使用最後一欄覆蓋
                    receipt_date_col = 24
                    worksheet.update_cell(1, receipt_date_col, '驗收日期')
                    print(f"DEBUG: 使用最後一欄第 {receipt_date_col} 欄作為驗收日期欄位")
        
        # 取得當前日期
        current_date = datetime.now().strftime('%Y%m%d')
        
        print(f"DEBUG: 更新欄位 - 驗收單驗收狀態: {receipt_status_col}, 驗收人員: {receipt_person_col}, 驗收日期: {receipt_date_col}")
        
        # 更新驗收單驗收狀態
        worksheet.update_cell(row_index, receipt_status_col, receipt_status)
        
        # 更新驗收人員（自動帶入登入者姓名）
        worksheet.update_cell(row_index, receipt_person_col, current_user_name)
        
        # 更新驗收日期（自動帶入當前日期）
        worksheet.update_cell(row_index, receipt_date_col, current_date)
        
        print(f"DEBUG: 已更新 - 驗收單驗收狀態: {receipt_status}, 驗收人員: {current_user_name}, 驗收日期: {current_date}")
        
        # 新增：驗收單驗收狀態為「已驗收」時，呼叫 GAS 進行 Google Sheets 單列保護
        print(f"DEBUG: receipt_status 判斷值: [{receipt_status}]")
        if receipt_status.strip() == '已驗收':
            print("DEBUG: 準備呼叫 GAS set_purchase_protection")
            if os.getenv('ENABLE_GAS_PROTECTION', 'false').lower() == 'true':
                protection_result = set_purchase_protection(purchase_no, "驗收完成自動鎖定")
                print(f"GAS 保護結果: {protection_result}")

        return jsonify({
            'success': True, 
            'message': '驗收單驗收狀態已更新',
            'receipt_person': current_user_name,
            'receipt_date': current_date
        })
        
    except Exception as e:
        print(f"更新驗收單驗收狀態失敗: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/purchase-search')
def purchase_search():
    """總務事務用品搜尋頁面"""
    if 'logged_in' not in session or not session['logged_in']:
        return redirect(url_for('index'))
    
    return render_template('purchase_search.html')

@app.route('/purchase-no-search', methods=['GET', 'POST'])
def purchase_no_search():
    """依請購單號搜尋頁面"""
    if 'logged_in' not in session or not session['logged_in']:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        # 處理 POST 請求（表單提交）
        start_purchase_no = request.form.get('start_purchase_no', '').strip()
        end_purchase_no = request.form.get('end_purchase_no', '').strip()
        
        if not start_purchase_no or not end_purchase_no:
            flash('請輸入開始和結束請購單號', 'error')
            return render_template('purchase_no_search.html')
        
        # 重定向到搜尋結果頁面或使用 AJAX 處理
        return render_template('purchase_no_search.html', 
                             start_purchase_no=start_purchase_no,
                             end_purchase_no=end_purchase_no)
    
    return render_template('purchase_no_search.html')

@app.route('/purchase-date-search', methods=['GET', 'POST'])
def purchase_date_search():
    """依請購日期搜尋頁面"""
    if 'logged_in' not in session or not session['logged_in']:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        # 處理 POST 請求（表單提交）
        start_date = request.form.get('start_date', '').strip()
        end_date = request.form.get('end_date', '').strip()
        
        if not start_date or not end_date:
            flash('請選擇開始和結束日期', 'error')
            return render_template('purchase_date_search.html')
        
        # 重定向到搜尋結果頁面或使用 AJAX 處理
        return render_template('purchase_date_search.html', 
                             start_date=start_date,
                             end_date=end_date)
    
    return render_template('purchase_date_search.html')

@app.route('/purchase-department-search', methods=['GET', 'POST'])
def purchase_department_search():
    """依請購部門搜尋頁面"""
    if 'logged_in' not in session or not session['logged_in']:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        # 處理 POST 請求（表單提交）
        department = request.form.get('department', '').strip()
        
        if not department:
            flash('請選擇請購部門', 'error')
            return render_template('purchase_department_search.html')
        
        # 重定向到搜尋結果頁面或使用 AJAX 處理
        return render_template('purchase_department_search.html', 
                             department=department)
    
    return render_template('purchase_department_search.html')

@app.route('/purchase-approval-status-search', methods=['GET', 'POST'])
def purchase_approval_status_search():
    """依簽核狀態搜尋頁面"""
    if 'logged_in' not in session or not session['logged_in']:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        # 處理 POST 請求（表單提交）
        approval_type = request.form.get('approval_type', '').strip()
        approval_status = request.form.get('approval_status', '').strip()
        
        if not approval_type or not approval_status:
            flash('請選擇簽核類型和狀態', 'error')
            return render_template('purchase_approval_status_search.html')
        
        # 重定向到搜尋結果頁面或使用 AJAX 處理
        return render_template('purchase_approval_status_search.html', 
                             approval_type=approval_type,
                             approval_status=approval_status)
    
    return render_template('purchase_approval_status_search.html')

@app.route('/test_search_simple.html')
def test_search_simple():
    """簡單搜尋測試頁面"""
    return render_template('test_search_simple.html')

@app.route('/test_login.html')
def test_login():
    """登入測試頁面"""
    return render_template('test_login.html')

@app.route('/test_purchase_no_search.html')
def test_purchase_no_search():
    """測試請購單號搜尋頁面"""
    return render_template('test_purchase_no_search.html')

@app.route('/test_approval_search.html')
def test_approval_search():
    """測試簽核狀態搜尋頁面"""
    return render_template('test_approval_search.html')

@app.route('/debug_approval_data.html')
def debug_approval_data():
    """調試簽核狀態資料頁面"""
    return render_template('debug_approval_data.html')

@app.route('/test_receipt_status.html')
def test_receipt_status():
    """測試驗收狀態更新頁面"""
    return render_template('test_receipt_status.html')

@app.route('/test_receipt_person.html')
def test_receipt_person():
    """測試驗收人員自動帶入頁面"""
    return render_template('test_receipt_person.html')

@app.route('/test_receipt_approval.html')
def test_receipt_approval():
    """測試驗收簽核人員自動帶入頁面"""
    return render_template('test_receipt_approval.html')

@app.route('/test_receipt_approval_check.html')
def test_receipt_approval_check():
    """測試驗收單檢核機制頁面"""
    return render_template('test_receipt_approval_check.html')

@app.route('/debug-receipt-data')
def debug_receipt_data():
    """調試驗收單資料"""
    try:
        client = get_google_sheets_client()
        spreadsheet_id = os.getenv('SPREADSHEET_ID')
        spreadsheet = client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet('請購單')
        all_records = get_safe_records(worksheet)
        
        # 分析所有記錄的簽核狀態
        approval_analysis = {}
        total_records = len(all_records)
        approved_count = 0
        
        for record in all_records:
            purchase_no = record.get('請購單號', 'N/A')
            
            # 檢查多個可能的簽核欄位名稱
            approval_status = (
                record.get('請購單簽核') or 
                record.get('簽核狀態') or 
                record.get('approval_status') or 
                record.get('狀態') or 
                '未設定'
            )
            
            if approval_status in ['核准', 'approved', 'APPROVED', 'Approved', '已核准', '已批准']:
                approved_count += 1
            
            approval_analysis[purchase_no] = {
                'approval_status': approval_status,
                'department': record.get('請購部門'),
                'applicant': record.get('申請人'),
                'item_name': record.get('品名'),
                'receipt_status': record.get('驗收單狀態'),
                'receipt_approval_status': record.get('驗收簽核狀態')
            }
        
        return jsonify({
            'success': True,
            'total_records': total_records,
            'approved_count': approved_count,
            'approval_analysis': approval_analysis,
            'columns': list(all_records[0].keys()) if all_records else []
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        })

@app.route('/debug-receipt.html')
def debug_receipt_page():
    """調試驗收單資料頁面"""
    return render_template('debug_receipt.html')

@app.route('/test-receipt-display.html')
def test_receipt_display():
    """測試驗收單顯示邏輯頁面"""
    return render_template('test_receipt_display.html')

@app.route('/get-purchase-status/<purchase_no>')
def get_purchase_status(purchase_no):
    """獲取請購單的當前狀態"""
    if 'logged_in' not in session or not session['logged_in']:
        return jsonify({'success': False, 'message': '未登入'})
    
    try:
        client = get_google_sheets_client()
        spreadsheet_id = os.getenv('SPREADSHEET_ID')
        spreadsheet = client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet('請購單')
        
        # 找到對應的請購單號
        all_records = get_safe_records(worksheet)
        purchase_record = None
        
        for record in all_records:
            if str(record.get('請購單號', '')).replace('-', '') == purchase_no:
                purchase_record = record
                break
        
        if not purchase_record:
            return jsonify({'success': False, 'message': '找不到請購單'})
        
        # 取得驗收狀態（嘗試多個可能的欄位名稱）
        receipt_status = (
            purchase_record.get('驗收單狀態') or 
            purchase_record.get('驗收狀態') or 
            purchase_record.get('receipt_status') or 
            '待驗收'
        )
        
        return jsonify({
            'success': True,
            'data': {
                'purchase_no': purchase_record.get('請購單號', ''),
                'receipt_status': receipt_status,
                'approval_status': purchase_record.get('請購單簽核', ''),
                'receipt_approval_status': purchase_record.get('驗收簽核狀態', ''),
                'receipt_approval_person': purchase_record.get('驗收簽核人員', ''),
                'receipt_approval_date': purchase_record.get('驗收簽核日期', '')
            }
        })
        
    except Exception as e:
        print(f"獲取請購單狀態失敗: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/debug-data')
def debug_data():
    """調試資料頁面"""
    try:
        client = get_google_sheets_client()
        spreadsheet_id = os.getenv('SPREADSHEET_ID')
        spreadsheet = client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet('請購單')
        all_records = get_safe_records(worksheet)
        
        # 只返回前5筆記錄用於調試
        debug_records = all_records[:5] if len(all_records) > 5 else all_records
        
        # 檢查簽核狀態欄位
        approval_statuses = {}
        for record in all_records[:10]:  # 檢查前10筆記錄
            purchase_no = record.get('請購單號', 'N/A')
            approval_statuses[purchase_no] = {
                '簽核': record.get('請購單簽核'),
                '簽核狀態': record.get('簽核狀態'),
                'approval_status': record.get('approval_status'),
                '狀態': record.get('狀態')
            }
        
        # 檢查特定請購單的驗收狀態欄位
        target_purchase_no = '20250718-001'
        target_record = None
        for record in all_records:
            if record.get('請購單號') == target_purchase_no:
                target_record = record
                break
        
        return jsonify({
            'success': True,
            'total_records': len(all_records),
            'sample_records': debug_records,
            'columns': list(all_records[0].keys()) if all_records else [],
            'approval_statuses': approval_statuses,
            'target_purchase': target_record,
            'target_receipt_status_fields': {
                '驗收單狀態': target_record.get('驗收單狀態') if target_record else None,
                '驗收狀態': target_record.get('驗收狀態') if target_record else None,
                'receipt_status': target_record.get('receipt_status') if target_record else None
            } if target_record else None
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        })

@app.route('/search-purchase-requests', methods=['POST'])
def search_purchase_requests():
    """搜尋請購單"""
    if 'logged_in' not in session or not session['logged_in']:
        return jsonify({'success': False, 'message': '未登入'})
    
    try:
        data = request.get_json()
        search_type = data.get('search_type')
        
        print(f"DEBUG: 搜尋類型: {search_type}")
        print(f"DEBUG: 搜尋資料: {data}")
        
        # 取得所有請購單資料
        client = get_google_sheets_client()
        if not client:
            return jsonify({'success': False, 'message': '無法建立 Google Sheets 客戶端'})
        spreadsheet_id = os.getenv('SPREADSHEET_ID')
        if not spreadsheet_id:
            return jsonify({'success': False, 'message': '未設定試算表 ID'})
        spreadsheet = client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet('請購單')
        all_records = get_safe_records(worksheet)
        
        print(f"DEBUG: 總記錄數: {len(all_records)}")
        if all_records:
            print(f"DEBUG: 第一筆記錄: {all_records[0]}")
        
        # 根據搜尋類型進行篩選
        filtered_records = []
        
        for record in all_records:
            if search_type == 'purchase_no':
                # 依請購單號搜尋
                purchase_no = data.get('purchase_no', '').strip()
                record_purchase_no = str(record.get('請購單號', '')).strip()
                if purchase_no and record_purchase_no.find(purchase_no) != -1:
                    filtered_records.append(record)
            elif search_type == 'purchase_no_range':
                # 依請購單號範圍搜尋
                start_purchase_no = data.get('start_purchase_no', '').strip()
                end_purchase_no = data.get('end_purchase_no', '').strip()
                record_purchase_no = str(record.get('請購單號', '')).strip()
                
                print(f"DEBUG: 搜尋範圍 {start_purchase_no} 到 {end_purchase_no}, 記錄請購單號: {record_purchase_no}")
                
                # 如果沒有請購單號資料，跳過
                if not record_purchase_no:
                    print(f"DEBUG: 跳過空請購單號記錄")
                    continue
                
                # 如果沒有輸入範圍，返回所有記錄
                if not start_purchase_no and not end_purchase_no:
                    print(f"DEBUG: 沒有輸入範圍，返回所有記錄")
                    filtered_records.append(record)
                    continue
                
                # 如果只有開始號碼，搜尋大於等於開始號碼的記錄
                if start_purchase_no and not end_purchase_no:
                    if record_purchase_no >= start_purchase_no:
                        print(f"DEBUG: 符合開始號碼條件，加入結果: {record_purchase_no}")
                        filtered_records.append(record)
                    continue
                
                # 如果只有結束號碼，搜尋小於等於結束號碼的記錄
                if not start_purchase_no and end_purchase_no:
                    if record_purchase_no <= end_purchase_no:
                        print(f"DEBUG: 符合結束號碼條件，加入結果: {record_purchase_no}")
                        filtered_records.append(record)
                    continue
                
                # 如果有開始和結束號碼，進行範圍搜尋
                if start_purchase_no and end_purchase_no:
                    # 嘗試數字比較（如果請購單號是純數字）
                    try:
                        start_num = int(start_purchase_no.replace('-', ''))
                        end_num = int(end_purchase_no.replace('-', ''))
                        record_num = int(record_purchase_no.replace('-', ''))
                        
                        if start_num <= record_num <= end_num:
                            print(f"DEBUG: 數字比較符合條件，加入結果: {record_purchase_no}")
                            filtered_records.append(record)
                    except:
                        # 如果數字比較失敗，使用字串比較
                        if start_purchase_no <= record_purchase_no <= end_purchase_no:
                            print(f"DEBUG: 字串比較符合條件，加入結果: {record_purchase_no}")
                            filtered_records.append(record)
                    
            elif search_type == 'create_date':
                # 依請購日期搜尋
                start_date = data.get('start_date', '')
                end_date = data.get('end_date', '')
                record_date = str(record.get('請購日期', ''))
                
                if start_date and end_date:
                    # 轉換日期格式進行比較
                    try:
                        from datetime import datetime
                        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                        
                        if record_date:
                            # 處理不同的日期格式
                            record_date_clean = record_date.replace('-', '').replace('/', '')
                            if len(record_date_clean) == 8:  # YYYYMMDD
                                record_dt = datetime.strptime(record_date_clean, '%Y%m%d')
                            elif len(record_date_clean) == 10:  # YYYY-MM-DD (舊格式)
                                record_dt = datetime.strptime(record_date_clean, '%Y-%m-%d')
                            else:
                                continue
                            
                            if start_dt <= record_dt <= end_dt:
                                filtered_records.append(record)
                    except:
                        continue
                        
            elif search_type == 'department':
                # 依請購部門搜尋
                department = data.get('department', '').strip()
                if department and record.get('請購部門', '') == department:
                    filtered_records.append(record)
                    
            elif search_type == 'applicant':
                # 依申請人搜尋
                applicant = data.get('applicant', '').strip()
                record_applicant = str(record.get('申請人', '')).strip()
                if applicant and record_applicant.find(applicant) != -1:
                    filtered_records.append(record)
                    
            elif search_type == 'approval_status':
                # 依簽核狀態搜尋
                approval_type = data.get('approval_type', '').strip()
                approval_status = data.get('approval_status', '').strip()
                
                print(f"DEBUG: 簽核狀態搜尋 - 類型: {approval_type}, 狀態: {approval_status}")
                
                # 嘗試多個可能的欄位名稱
                purchase_approval_fields = ['請購單簽核', '簽核', '簽核狀態', 'approval_status']
                receipt_approval_fields = ['驗收簽核人員', '驗收簽核日期', '驗收單簽核狀態', '驗收簽核', 'receipt_approval_status']
                
                if approval_type == 'purchase_approval':
                    # 請購單簽核狀態搜尋
                    record_approval = None
                    for field in purchase_approval_fields:
                        if field in record:
                            record_approval = str(record.get(field, '')).strip()
                            print(f"DEBUG: 找到請購單簽核欄位 '{field}': '{record_approval}'")
                            break
                    
                    if record_approval is None:
                        print(f"DEBUG: 未找到請購單簽核欄位，記錄欄位: {list(record.keys())}")
                        continue
                    
                    print(f"DEBUG: 請購單簽核比較 - 搜尋: '{approval_status}' vs 記錄: '{record_approval}'")
                    
                    # 更靈活的匹配邏輯
                    if approval_status and (record_approval == approval_status or 
                                          record_approval.find(approval_status) != -1 or
                                          approval_status.find(record_approval) != -1):
                        print(f"DEBUG: 請購單簽核狀態匹配，加入結果")
                        filtered_records.append(record)
                        
                elif approval_type == 'receipt_approval':
                    # 驗收單簽核狀態搜尋 - 根據驗收簽核人員判斷狀態
                    receipt_approver = str(record.get('驗收簽核人員', '')).strip()
                    receipt_approval_date = str(record.get('驗收簽核日期', '')).strip()
                    
                    print(f"DEBUG: 驗收簽核人員: '{receipt_approver}', 驗收簽核日期: '{receipt_approval_date}'")
                    
                    # 根據驗收簽核狀態進行判斷
                    if approval_status == '待簽核':
                        # 待簽核：沒有驗收簽核人員或驗收簽核日期
                        if not receipt_approver and not receipt_approval_date:
                            print(f"DEBUG: 驗收單待簽核狀態匹配，加入結果")
                            filtered_records.append(record)
                    elif approval_status == '核准':
                        # 核准：有驗收簽核人員和驗收簽核日期
                        if receipt_approver and receipt_approval_date:
                            print(f"DEBUG: 驗收單核准狀態匹配，加入結果")
                            filtered_records.append(record)
                    elif approval_status == '駁回':
                        # 駁回：有驗收簽核人員但沒有驗收簽核日期（或其他邏輯）
                        if receipt_approver and not receipt_approval_date:
                            print(f"DEBUG: 驗收單駁回狀態匹配，加入結果")
                            filtered_records.append(record)
                    
            elif search_type == 'custom':
                # 自訂搜尋 - 組合多個條件
                matches = True
                
                # 請購單號
                custom_purchase_no = data.get('custom_purchase_no', '').strip()
                record_purchase_no = str(record.get('請購單號', '')).strip()
                if custom_purchase_no and record_purchase_no.find(custom_purchase_no) == -1:
                    matches = False
                
                # 請購部門
                custom_department = data.get('custom_department', '').strip()
                record_department = str(record.get('請購部門', '')).strip()
                if custom_department and record_department != custom_department:
                    matches = False
                
                # 申請人
                custom_applicant = data.get('custom_applicant', '').strip()
                record_applicant = str(record.get('申請人', '')).strip()
                if custom_applicant and record_applicant.find(custom_applicant) == -1:
                    matches = False
                
                # 簽核狀態
                custom_approval_status = data.get('custom_approval_status', '').strip()
                if custom_approval_status and record.get('請購單簽核', '') != custom_approval_status:
                    matches = False
                
                # 日期範圍
                custom_start_date = data.get('custom_start_date', '')
                custom_end_date = data.get('custom_end_date', '')
                if custom_start_date and custom_end_date:
                    try:
                        from datetime import datetime
                        start_dt = datetime.strptime(custom_start_date, '%Y-%m-%d')
                        end_dt = datetime.strptime(custom_end_date, '%Y-%m-%d')
                        
                        record_date = str(record.get('請購日期', ''))
                        if record_date:
                            record_date_clean = record_date.replace('-', '').replace('/', '')
                            if len(record_date_clean) == 8:  # YYYYMMDD
                                record_dt = datetime.strptime(record_date_clean, '%Y%m%d')
                            elif len(record_date_clean) == 10:  # YYYY-MM-DD (舊格式)
                                record_dt = datetime.strptime(record_date_clean, '%Y-%m-%d')
                            else:
                                matches = False
                            
                            if not (start_dt <= record_dt <= end_dt):
                                matches = False
                    except:
                        matches = False
                
                if matches:
                    filtered_records.append(record)
        
        print(f"DEBUG: 篩選後記錄數: {len(filtered_records)}")
        
        return jsonify({
            'success': True, 
            'results': filtered_records,
            'total_count': len(filtered_records)
        })
        
    except Exception as e:
        print(f"搜尋請購單失敗: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/test-search-purchase-requests', methods=['POST'])
def test_search_purchase_requests():
    """測試用搜尋請購單（不需要登入）"""
    # 直接調用原始搜尋邏輯，跳過登入檢查
    try:
        data = request.get_json()
        search_type = data.get('search_type')
        
        print(f"TEST DEBUG: 搜尋類型: {search_type}")
        print(f"TEST DEBUG: 搜尋資料: {data}")
        
        # 取得所有請購單資料
        client = get_google_sheets_client()
        if not client:
            return jsonify({'success': False, 'message': '無法建立 Google Sheets 客戶端'})
        spreadsheet_id = os.getenv('SPREADSHEET_ID')
        if not spreadsheet_id:
            return jsonify({'success': False, 'message': '未設定試算表 ID'})
        spreadsheet = client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet('請購單')
        all_records = get_safe_records(worksheet)
        
        print(f"TEST DEBUG: 總記錄數: {len(all_records)}")
        if all_records:
            print(f"TEST DEBUG: 第一筆記錄: {all_records[0]}")
        
        # 根據搜尋類型進行篩選
        filtered_records = []
        
        for record in all_records:
            if search_type == 'purchase_no_range':
                # 依請購單號範圍搜尋
                start_purchase_no = data.get('start_purchase_no', '').strip()
                end_purchase_no = data.get('end_purchase_no', '').strip()
                record_purchase_no = str(record.get('請購單號', '')).strip()
                
                print(f"TEST DEBUG: 搜尋範圍 {start_purchase_no} 到 {end_purchase_no}, 記錄請購單號: {record_purchase_no}")
                
                # 如果沒有請購單號資料，跳過
                if not record_purchase_no:
                    print(f"TEST DEBUG: 跳過空請購單號記錄")
                    continue
                
                # 如果沒有輸入範圍，返回所有記錄
                if not start_purchase_no and not end_purchase_no:
                    print(f"TEST DEBUG: 沒有輸入範圍，返回所有記錄")
                    filtered_records.append(record)
                    continue
                
                # 如果只有開始號碼，搜尋大於等於開始號碼的記錄
                if start_purchase_no and not end_purchase_no:
                    if record_purchase_no >= start_purchase_no:
                        print(f"TEST DEBUG: 符合開始號碼條件，加入結果: {record_purchase_no}")
                        filtered_records.append(record)
                    continue
                
                # 如果只有結束號碼，搜尋小於等於結束號碼的記錄
                if not start_purchase_no and end_purchase_no:
                    if record_purchase_no <= end_purchase_no:
                        print(f"TEST DEBUG: 符合結束號碼條件，加入結果: {record_purchase_no}")
                        filtered_records.append(record)
                    continue
                
                # 如果有開始和結束號碼，進行範圍搜尋
                if start_purchase_no and end_purchase_no:
                    # 嘗試數字比較（如果請購單號是純數字）
                    try:
                        start_num = int(start_purchase_no.replace('-', ''))
                        end_num = int(end_purchase_no.replace('-', ''))
                        record_num = int(record_purchase_no.replace('-', ''))
                        
                        if start_num <= record_num <= end_num:
                            print(f"TEST DEBUG: 數字比較符合條件，加入結果: {record_purchase_no}")
                            filtered_records.append(record)
                    except:
                        # 如果數字比較失敗，使用字串比較
                        if start_purchase_no <= record_purchase_no <= end_purchase_no:
                            print(f"TEST DEBUG: 字串比較符合條件，加入結果: {record_purchase_no}")
                            filtered_records.append(record)
        
        print(f"TEST DEBUG: 篩選後記錄數: {len(filtered_records)}")
        
        return jsonify({
            'success': True, 
            'results': filtered_records,
            'total_count': len(filtered_records)
        })
        
    except Exception as e:
        print(f"測試搜尋請購單失敗: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/purchase-request/new', methods=['GET', 'POST'])
def purchase_request_new():
    if 'logged_in' not in session or not session['logged_in']:
        return redirect(url_for('index'))
    username = session.get('username')
    user_info = get_user_info(username)
    if request.method == 'POST':
        # 強制由系統產生唯一請購單號
        purchase_no = generate_purchase_no()
        today = datetime.now().strftime('%Y%m%d')
        department = request.form.get('department')
        applicant = user_info['name']
        applicant_mail = user_info['mail']
        item_name = request.form.get('item_name')
        spec = request.form.get('spec')
        quantity = request.form.get('quantity')
        unit = request.form.get('unit')
        need_date = request.form.get('need_date')
        purpose = request.form.get('purpose')
        note = request.form.get('note')
        sign_status = request.form.get('sign_status')
        reject_reason = request.form.get('reject_reason')
        attachment_url = ''
        if 'attachment' in request.files:
            file = request.files['attachment']
            if file and file.filename:
                folder_id = os.getenv('PURCHASE_REQUEST_ATTACHMENT_FOLDER_ID') # 從環境變數讀取資料夾ID
                if not folder_id:
                    flash('未設定請購單附件資料夾ID', 'error')
                    return redirect(url_for('purchase_request_new'))
                attachment_url = upload_to_drive(file, folder_id)
                print(f'=== 附件上傳結果 ===')
                print(f'檔案名稱: {file.filename}')
                print(f'附件URL: {attachment_url}')
                print(f'附件URL長度: {len(attachment_url) if attachment_url else 0}')
        try:
            client = get_google_sheets_client()
            spreadsheet_id = os.getenv('SPREADSHEET_ID')
            spreadsheet = client.open_by_key(spreadsheet_id)
            worksheet = spreadsheet.worksheet('請購單')
            
            # 準備要寫入的資料 - 按照 Google Sheets 的欄位順序
            row_data = [
                purchase_no,        # 1. 請購單號
                today,             # 2. 請購日期
                department,        # 3. 請購部門
                applicant,         # 4. 申請人
                applicant_mail,    # 5. mail
                item_name,         # 6. 品名
                spec,              # 7. 規格
                quantity,          # 8. 數量
                unit,              # 9. 單位
                need_date,         # 10. 需求日期
                purpose,           # 11. 用途
                attachment_url,    # 12. 上傳附件
                note,              # 13. 備註
                sign_status,       # 14. 請購單簽核
                reject_reason      # 15. 請購單駁回原因
            ]
            
            print(f'=== 寫入Google Sheets ===')
            print(f'請購單號: {purchase_no}')
            print(f'附件URL: {attachment_url}')
            print(f'資料列: {row_data}')
            
            worksheet.append_row(row_data)
            # 寫入日誌
            user_info = get_user_info(session['username'])
            now = datetime.now().strftime('%Y%m%d %H:%M')
            update_system_log(user_info.get('name', session['username']), action_str=f'請購單建立 {purchase_no} {now}')
            flash('請購單已成功建立！', 'success')
            return redirect(url_for('purchase_request_new'))
        except Exception as e:
            print(f"寫入請購單失敗: {e}")
            flash('請購單建立失敗，請稍後再試', 'error')
            return redirect(url_for('purchase_request_new'))
    purchase_no = generate_purchase_no()
    today = datetime.now().strftime('%Y%m%d')
    departments = get_dropdown_list('請購部門', '部門名稱')
    units = get_dropdown_list('單位', '單位名稱')
    return render_template('purchase_request_new.html',
        purchase_no=purchase_no,
        today=today,
        departments=departments,
        units=units,
        username=user_info['name'],
        user_mail=user_info['mail']
    )

@app.route('/check-sheet-protection')
def check_sheet_protection():
    """檢查 Google Sheets 保護狀態"""
    if 'logged_in' not in session or not session['logged_in']:
        return jsonify({'success': False, 'message': '未登入'})
    
    try:
        client = get_google_sheets_client()
        spreadsheet_id = os.getenv('SPREADSHEET_ID')
        
        # 使用 Google Sheets API 檢查保護狀態
        from googleapiclient.discovery import build
        from googleapiclient.errors import HttpError
        
        # 取得認證憑證
        creds = None
        if os.path.exists('token.json'):
            from google.auth.transport.requests import Request
            from google.oauth2.credentials import Credentials
            creds = Credentials.from_authorized_user_file('token.json', ['https://www.googleapis.com/auth/spreadsheets'])
        elif os.path.exists('service-account-key.json'):
            from google.oauth2 import service_account
            creds = service_account.Credentials.from_service_account_file(
                'service-account-key.json',
                scopes=['https://www.googleapis.com/auth/spreadsheets']
            )
        
        if creds and creds.valid:
            service = build('sheets', 'v4', credentials=creds)
            
            # 取得試算表資訊
            spreadsheet = service.spreadsheets().get(
                spreadsheetId=spreadsheet_id,
                ranges=[],
                includeGridData=False
            ).execute()
            
            protection_info = []
            if 'sheets' in spreadsheet:
                for sheet in spreadsheet['sheets']:
                    sheet_name = sheet['properties']['title']
                    if 'protectedRanges' in sheet:
                        for protected_range in sheet['protectedRanges']:
                            protection_info.append({
                                'sheet_name': sheet_name,
                                'protected_range_id': protected_range['protectedRangeId'],
                                'description': protected_range.get('description', '無描述'),
                                'range': protected_range['range']
                            })
            
            return jsonify({
                'success': True,
                'protection_info': protection_info,
                'total_protected_ranges': len(protection_info)
            })
        else:
            return jsonify({
                'success': False,
                'message': '無法取得有效的認證憑證'
            })
            
    except Exception as e:
        print(f"檢查保護狀態失敗: {e}")
        return jsonify({'error': str(e)}), 500

def write_system_log(name, login_time=None, logout_time=None, action=''):
    try:
        client = get_google_sheets_client()
        spreadsheet_id = os.getenv('SPREADSHEET_ID')
        spreadsheet = client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet('系統日誌')
        row = [
            name,
            login_time or '',
            logout_time or '',
            action
        ]
        worksheet.append_row(row)
    except Exception as e:
        print(f"寫入系統日誌失敗: {e}")

def update_system_log(name, action_str=None, logout_time=None):
    try:
        client = get_google_sheets_client()
        spreadsheet_id = os.getenv('SPREADSHEET_ID')
        spreadsheet = client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet('系統日誌')
        records = worksheet.get_all_records()
        # 找到該用戶最後一筆尚未登出的紀錄
        target_row = None
        for idx, row in reversed(list(enumerate(records, start=2))):  # start=2 因為第1行是標題
            print(f"檢查日誌: {row.get('姓名')} 登出時間: {row.get('登出時間')}")
            if row.get('姓名') == name and not row.get('登出時間'):
                target_row = idx
                break
        print(f"找到要更新的 row: {target_row}")
        if target_row:
            if action_str:
                old_content = worksheet.cell(target_row, 4).value or ''
                new_content = (old_content + action_str + ' ; ').strip()
                worksheet.update_cell(target_row, 4, new_content)
            if logout_time:
                worksheet.update_cell(target_row, 3, logout_time)
        else:
            print(f"找不到未登出的日誌列，無法寫入登出時間")
    except Exception as e:
        print(f'更新系統日誌失敗: {e}')

@app.route('/user-management', methods=['GET'])
def user_management():
    client = get_google_sheets_client()
    spreadsheet_id = os.getenv('SPREADSHEET_ID')
    spreadsheet = client.open_by_key(spreadsheet_id)
    worksheet = spreadsheet.worksheet('使用者帳號')
    users = worksheet.get_all_records()
    return render_template('user_management.html', users=users)

@app.route('/user-management/add', methods=['POST'])
def user_management_add():
    name = request.form.get('name')
    username = request.form.get('username')
    password = request.form.get('password')
    role = request.form.get('role')
    mail = request.form.get('mail')
    client = get_google_sheets_client()
    spreadsheet_id = os.getenv('SPREADSHEET_ID')
    spreadsheet = client.open_by_key(spreadsheet_id)
    worksheet = spreadsheet.worksheet('使用者帳號')
    worksheet.append_row([name, username, password, role, mail])
    flash('新增成功', 'success')
    return redirect(url_for('user_management'))

@app.route('/user-management/update', methods=['POST'])
def user_management_update():
    username = request.form.get('username')
    name = request.form.get('name')
    password = request.form.get('password')
    role = request.form.get('role')
    mail = request.form.get('mail')
    client = get_google_sheets_client()
    spreadsheet_id = os.getenv('SPREADSHEET_ID')
    spreadsheet = client.open_by_key(spreadsheet_id)
    worksheet = spreadsheet.worksheet('使用者帳號')
    records = worksheet.get_all_records()
    for idx, row in enumerate(records, start=2):
        if row.get('帳號') == username:
            worksheet.update(f'A{idx}:E{idx}', [[name, username, password, role, mail]])
            break
    flash('更新成功', 'success')
    return redirect(url_for('user_management'))

@app.route('/user-management/delete', methods=['POST'])
def user_management_delete():
    username = request.form.get('username')
    client = get_google_sheets_client()
    spreadsheet_id = os.getenv('SPREADSHEET_ID')
    spreadsheet = client.open_by_key(spreadsheet_id)
    worksheet = spreadsheet.worksheet('使用者帳號')
    records = worksheet.get_all_records()
    for idx, row in enumerate(records, start=2):
        if row.get('帳號') == username:
            worksheet.delete_rows(idx)
            break
    flash('刪除成功', 'success')
    return redirect(url_for('user_management'))

@app.route('/system-log')
def system_log():
    if 'logged_in' not in session or not session['logged_in']:
        return redirect(url_for('index'))
    if session.get('role') == '一般人員':
        return '權限不足，無法存取此頁面', 403
    client = get_google_sheets_client()
    spreadsheet_id = os.getenv('SPREADSHEET_ID')
    spreadsheet = client.open_by_key(spreadsheet_id)
    worksheet = spreadsheet.worksheet('系統日誌')
    records = worksheet.get_all_records()
    # 取得所有姓名與登入時間選項
    names = sorted(set(r.get('姓名','') for r in records if r.get('姓名')))
    login_times = sorted(set(r.get('登入時間','') for r in records if r.get('登入時間')))
    return render_template('system_log.html', records=records, names=names, login_times=login_times)

@app.route('/attendance-check')
def attendance_check():
    if 'logged_in' not in session or not session['logged_in']:
        return redirect(url_for('index'))
    # 假資料
    attendance_data = [
        {'name': '王小明', 'emp_id': 'A001', 'shift': '早班', 'clock_in': '08:01', 'clock_out': '17:00', 'status': '正常'},
        {'name': '李小華', 'emp_id': 'A002', 'shift': '晚班', 'clock_in': '16:59', 'clock_out': '01:00', 'status': '正常'},
        {'name': '張大同', 'emp_id': 'A003', 'shift': '早班', 'clock_in': '08:20', 'clock_out': '17:05', 'status': '異常'},
    ]
    return render_template('attendance_check.html', attendance_data=attendance_data)

def fetch_taiwan_holidays(year, api_key):
    calendar_id = 'zh.taiwan%23holiday%40group.v.calendar.google.com'
    time_min = f"{year}-01-01T00:00:00Z"
    time_max = f"{year}-12-31T23:59:59Z"
    url = (
        f"https://www.googleapis.com/calendar/v3/calendars/{calendar_id}/events"
        f"?key={api_key}&timeMin={time_min}&timeMax={time_max}&singleEvents=true&orderBy=startTime"
    )
    resp = requests.get(url)
    data = resp.json()
    holidays = {}
    for item in data.get('items', []):
        date = item['start'].get('date')
        summary = item.get('summary')
        holidays[date] = summary
    return holidays

def get_schedule_data(start_date=None, end_date=None):
    # 1. 取得 Google Drive 服務
    service_account_info = os.getenv('GOOGLE_SERVICE_ACCOUNT_INFO')
    if service_account_info:
        creds_dict = json.loads(service_account_info)
        creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
    else:
        creds = Credentials.from_service_account_file('service-account-key.json', scopes=SCOPES)
    drive_service = build('drive', 'v3', credentials=creds)
    # 2. 搜尋檔名包含「排班表」的 Excel 檔案
    results = drive_service.files().list(q="name contains '排班表' and mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' and trashed=false", fields="files(id, name)").execute()
    files = results.get('files', [])
    if not files:
        return []
    file_id = files[0]['id']
    # 3. 下載該 Excel 檔案
    request = drive_service.files().get_media(fileId=file_id)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        status, done = downloader.next_chunk()
    fh.seek(0)
    # 4. 用 pandas 讀取分頁「班表」
    try:
        df_raw = pd.read_excel(fh, sheet_name='班表', header=None)
    except Exception as e:
        return []
    df_raw.iloc[0] = df_raw.iloc[0].ffill()
    # 1. 找出所有「班別」欄的原始 index
    班別_cols = [i for i in range(2, df_raw.shape[1]) if str(df_raw.iloc[1, i]).strip() == '班別']
    # 先將第一列全部轉成 datetime 物件
    def parse_date(s):
        if isinstance(s, datetime):
            return s
        s = str(s).replace('-', '/')
        parts = s.split('/')
        if len(parts) == 3:
            return datetime(int(parts[0]), int(parts[1]), int(parts[2]))
        return None
    date_list = [parse_date(df_raw.iloc[0, i]) for i in 班別_cols]
    # 日期篩選
    if start_date and end_date:
        start_dt = parse_date(start_date)
        end_dt = parse_date(end_date)
        keep_idx = [i for i, d in enumerate(date_list) if d and start_dt <= d <= end_dt]
    else:
        keep_idx = list(range(len(班別_cols)))
    # 只保留 header、班別_cols、keep_cols、df
    班別_cols_filtered = [班別_cols[i] for i in keep_idx]
    keep_cols = [0, 1] + 班別_cols_filtered
    df = df_raw.iloc[:, keep_cols]
    # header
    def format_date(val):
        if isinstance(val, datetime):
            return f"{val.year}/{val.month}/{val.day}"
        s = str(val).strip()
        for fmt in ('%Y/%m/%d', '%Y-%m-%d', '%Y.%m.%d', '%Y%m%d', '%m/%d', '%m-%d'):
            try:
                d = datetime.strptime(s, fmt)
                if fmt in ('%m/%d', '%m-%d'):
                    d = d.replace(year=datetime.now().year)
                return f"{d.year}/{d.month}/{d.day}"
            except:
                continue
        return s
    header = ['員工編號', '姓名'] + [format_date(df_raw.iloc[0, i]) for i in 班別_cols_filtered]
    # 取得假日資料
    api_key = 'YOUR_GOOGLE_API_KEY'  # <--- 請填入你的 API Key
    year = int(header[2].split('/')[0])
    holidays = fetch_taiwan_holidays(year, api_key)  # dict: yyyy-mm-dd -> 假日名稱
    # 產生 header 標記，並記錄週六/週日欄位index
    header_display = header[:2]
    sat_cols = []  # 週六index（相對於header, data）
    sun_cols = []  # 週日index
    for idx, h in enumerate(header[2:]):
        parts = h.split('/')
        date_str = f"{parts[0]}-{int(parts[1]):02d}-{int(parts[2]):02d}"
        try:
            dt = datetime.strptime(date_str, '%Y-%m-%d')
        except:
            header_display.append(h)
            continue
        # 國定假日
        if date_str in holidays:
            if '補行上班' in holidays[date_str]:
                header_display.append(h)  # 只保留日期格式
            else:
                header_display.append(h)
        elif dt.weekday() == 5:
            header_display.append(h)
            sat_cols.append(idx+2)  # +2是因為header前面有兩欄
        elif dt.weekday() == 6:
            header_display.append(h)
            sun_cols.append(idx+2)
        else:
            header_display.append(h)
    header = header_display
    # 4. 資料列
    data = df.iloc[2:, :].values.tolist()
    # 刪除最後三列
    if len(data) > 3:
        data = data[:-3]
    # 依條件轉換資料（除第一列/第一欄/第二欄外）
    def convert_cell(val):
        v = str(val).strip()
        if v == '早班':
            return '01'
        elif v == '早12':
            return '02'
        elif v == '守衛早':
            return '02A'
        elif v == '中班':
            return '03'
        elif v == '夜12':
            return '04'
        elif v == '守衛夜':
            return '04A'
        elif v == '夜班':
            return '05'
        else:
            return '06'
    for row in data:
        # 將員工編號補零為三碼
        try:
            emp_id = str(int(float(row[0]))).zfill(3)
        except:
            emp_id = str(row[0])
        row[0] = (emp_id, False)
        for i in range(2, len(row)):
            # 週六欄位改為"休"，週日欄位改為"例"
            if i in sat_cols:
                row[i] = ('休', False)
            elif i in sun_cols:
                row[i] = ('例', False)
            else:
                code = convert_cell(row[i])
                is_red = code != '06'
                row[i] = (code, is_red)
        # 姓名欄不標記顏色
        row[1] = (row[1], False)
    schedule_data = [header] + data
    print('班別_cols:', 班別_cols)
    print('header raw:', [df_raw.iloc[0, i] for i in 班別_cols])
    print('header:', header)
    return schedule_data

@app.route('/export-schedule')
def export_schedule():
    if 'logged_in' not in session or not session['logged_in']:
        return redirect(url_for('index'))
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    if start_date and end_date:
        schedule_data = get_schedule_data(start_date, end_date)
    else:
        schedule_data = []
    return render_template('export_schedule.html', schedule_data=schedule_data, request=request)

@app.route('/download-schedule')
def download_schedule():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    schedule_data = get_schedule_data(start_date, end_date)
    if not schedule_data:
        return '找不到檔名為「排班表」的 Excel 檔案或資料異常', 404
    wb = openpyxl.Workbook()
    ws = wb.active
    import re
    for idx, row in enumerate(schedule_data):
        excel_row = []
        for cell in row:
            val = cell[0] if isinstance(cell, tuple) else cell
            if idx == 0 and isinstance(val, str):
                val = re.sub(r'<.*?>', '', val)
            excel_row.append(val)
        ws.append(excel_row)
    # 自動調整欄寬
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='班表匯出.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000) 